#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
سكريبت تحديث بيانات المستخدمين من ملف Excel باستخدام Python
- يحدث المستخدمين الموجودين فقط
- Ext.NO → avaya_extension
- Mobile → phone_work (Work Phone Number)
"""

import os
import sys
import json
import re
from openpyxl import load_workbook
from typing import Dict, List, Optional, Tuple
import psycopg2
from psycopg2.extras import RealDictCursor
from psycopg2 import sql

# إعداد ترميز UTF-8 للطباعة
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

class UsersExcelUpdater:
    def __init__(self, db_config: Dict):
        """تهيئة السكريبت مع إعدادات قاعدة البيانات"""
        self.db_config = db_config
        self.conn = None
        self.cursor = None
        self.departments_cache = {}
        self.phone_types_cache = {}
        self.errors = []
        self.success_count = 0
        self.updated_count = 0
        self.skipped_count = 0
        
    def connect_db(self):
        """الاتصال بقاعدة البيانات"""
        try:
            self.conn = psycopg2.connect(
                host=self.db_config['host'],
                port=self.db_config.get('port', 5432),
                database=self.db_config['database'],
                user=self.db_config['user'],
                password=self.db_config.get('password', '')
            )
            self.cursor = self.conn.cursor(cursor_factory=RealDictCursor)
            print("[OK] تم الاتصال بقاعدة البيانات بنجاح\n")
        except Exception as e:
            print(f"[ERROR] خطأ في الاتصال بقاعدة البيانات: {e}\n")
            sys.exit(1)
    
    def close_db(self):
        """إغلاق الاتصال بقاعدة البيانات"""
        if self.cursor:
            self.cursor.close()
        if self.conn:
            self.conn.close()
    
    def load_reference_data(self):
        """تحميل البيانات المرجعية (الأقسام، أنواع الهواتف)"""
        try:
            # تحميل الأقسام
            self.cursor.execute("SELECT id, name, name_ar FROM departments")
            departments = self.cursor.fetchall()
            for dept in departments:
                name_lower = dept['name'].lower() if dept['name'] else ''
                self.departments_cache[name_lower] = dept
            
            # تحميل/إنشاء نوع الهاتف "work"
            self.cursor.execute("""
                SELECT id FROM phone_types WHERE slug = 'work'
            """)
            phone_type = self.cursor.fetchone()
            
            if not phone_type:
                self.cursor.execute("""
                    INSERT INTO phone_types (slug, name, name_ar, is_active, sort_order)
                    VALUES ('work', 'Work', 'عمل', true, 1)
                    RETURNING id
                """)
                phone_type = self.cursor.fetchone()
                self.conn.commit()
            
            self.phone_types_cache['work'] = phone_type['id']
            
            print(f"[OK] تم تحميل البيانات المرجعية:")
            print(f"   - الأقسام: {len(self.departments_cache)}")
            print(f"   - نوع الهاتف (عمل): موجود\n")
            
        except Exception as e:
            print(f"[ERROR] خطأ في تحميل البيانات المرجعية: {e}\n")
            sys.exit(1)
    
    def clean_text(self, text: Optional[str]) -> Optional[str]:
        """تنظيف النص من المسافات الزائدة"""
        if text is None:
            return None
        # تحويل إلى نص إذا كان رقم أو نوع آخر
        if not isinstance(text, str):
            text = str(text)
        text = text.strip()
        text = re.sub(r'\s+', ' ', text)
        return text if text else None
    
    def find_department(self, department_name: Optional[str]) -> Optional[Dict]:
        """البحث عن قسم بالاسم"""
        if not department_name:
            return None
        
        search_name = department_name.lower().strip()
        
        # البحث المباشر
        if search_name in self.departments_cache:
            return self.departments_cache[search_name]
        
        # البحث الجزئي
        for dept_name, dept in self.departments_cache.items():
            if search_name in dept_name or dept_name in search_name:
                return dept
        
        return None
    
    def find_user(self, email: Optional[str], employee_id: Optional[str], name: Optional[str]) -> Optional[Dict]:
        """البحث عن مستخدم بالإيميل أو Employee ID أو الاسم"""
        user = None
        
        try:
            # البحث بالإيميل
            if email:
                self.cursor.execute("SELECT * FROM users WHERE email = %s", (email,))
                user = self.cursor.fetchone()
                if user:
                    return dict(user)
            
            # البحث بـ Employee ID
            if not user and employee_id:
                self.cursor.execute("SELECT * FROM users WHERE employee_id = %s", (employee_id,))
                user = self.cursor.fetchone()
                if user:
                    return dict(user)
            
            # البحث بالاسم (مطابقة جزئية)
            if not user and name:
                self.cursor.execute("SELECT * FROM users WHERE name ILIKE %s LIMIT 1", (f"%{name}%",))
                user = self.cursor.fetchone()
                if user:
                    return dict(user)
        except Exception as e:
            # في حالة خطأ، عمل rollback
            try:
                self.conn.rollback()
            except:
                pass
        
        return None
    
    def clean_phone_number(self, phone: Optional[str]) -> Optional[str]:
        """تنظيف رقم الهاتف"""
        if not phone:
            return None
        # إزالة جميع الأحرف غير الرقمية والرمز +
        cleaned = re.sub(r'[^\d+]', '', str(phone))
        return cleaned if cleaned else None
    
    def extract_first_email(self, email_string: Optional[str]) -> Optional[str]:
        """استخراج أول إيميل من سلسلة إيميلات"""
        if not email_string:
            return None
        
        # تقسيم الإيميلات بالسطر الجديد أو المسافة
        emails = re.split(r'[\r\n\s]+', email_string.strip())
        for email in emails:
            email = email.strip().lower()
            if email and '@' in email:
                return email
        
        return None
    
    def add_employee_emails(self, user_id: int, email_string: str):
        """إضافة الإيميلات إلى جدول employee_emails"""
        if not email_string:
            return
        
        # تقسيم الإيميلات
        emails = re.split(r'[\r\n\s]+', email_string.strip())
        added_count = 0
        
        for email in emails:
            email = email.strip().lower()
            if not email or '@' not in email:
                continue
            
            # التحقق من صحة الإيميل
            if not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
                continue
            
            try:
                # التحقق من وجود الإيميل
                self.cursor.execute("""
                    SELECT id FROM employee_emails 
                    WHERE employee_id = %s AND email_address = %s
                """, (user_id, email))
                
                existing = self.cursor.fetchone()
                
                if not existing:
                    # تحديد إذا كان هذا الإيميل الأساسي
                    self.cursor.execute("""
                        SELECT COUNT(*) as count FROM employee_emails 
                        WHERE employee_id = %s AND is_primary = true
                    """, (user_id,))
                    count = self.cursor.fetchone()['count']
                    is_primary = count == 0
                    
                    self.cursor.execute("""
                        INSERT INTO employee_emails 
                        (employee_id, email_address, email_type, is_primary, is_active, notes, created_at, updated_at)
                        VALUES (%s, %s, 'work', %s, true, NULL, NOW(), NOW())
                    """, (user_id, email, is_primary))
                    
                    added_count += 1
            except Exception as e:
                # تجاهل الأخطاء (مثل الجدول غير موجود)
                pass
        
        if added_count > 0:
            self.conn.commit()
            print(f"   [OK] تم إضافة {added_count} إيميل إلى جدول employee_emails")
    
    def extract_first_phone(self, phone_string: Optional[str]) -> Optional[str]:
        """استخراج أول رقم هاتف من سلسلة أرقام (مفصولة بسطر جديد أو مسافة)"""
        if not phone_string:
            return None
        
        # تقسيم الأرقام بالسطر الجديد أو المسافة
        phones = re.split(r'[\r\n\s]+', str(phone_string).strip())
        for phone in phones:
            phone = self.clean_phone_number(phone)
            if phone and len(phone) >= 10:  # رقم هاتف صحيح
                return phone
        
        return None

    def update_work_phone(self, user_id: int, phone_number: str):
        """تحديث أو إضافة رقم هاتف العمل"""
        if not phone_number:
            return
        
        # استخراج أول رقم هاتف من السلسلة
        phone_number = self.extract_first_phone(phone_number)
        if not phone_number:
            return
        
        try:
            # البحث عن رقم هاتف عمل موجود
            self.cursor.execute("""
                SELECT up.id FROM user_phones up
                JOIN phone_types pt ON up.phone_type_id = pt.id
                WHERE up.user_id = %s AND pt.slug = 'work'
                LIMIT 1
            """, (user_id,))
            
            work_phone = self.cursor.fetchone()
            
            if not work_phone:
                # إنشاء رقم هاتف عمل جديد
                self.cursor.execute("""
                    INSERT INTO user_phones 
                    (user_id, phone_type_id, phone_number, is_primary, created_at, updated_at)
                    VALUES (%s, %s, %s, true, NOW(), NOW())
                """, (user_id, self.phone_types_cache['work'], phone_number))
                print(f"   [OK] رقم الهاتف (Mobile): {phone_number}")
                self.conn.commit()
            else:
                # تحديث الرقم الموجود
                self.cursor.execute("""
                    UPDATE user_phones 
                    SET phone_number = %s, updated_at = NOW()
                    WHERE id = %s
                """, (phone_number, work_phone['id']))
                print(f"   [OK] رقم الهاتف (Mobile): تم التحديث إلى {phone_number}")
                self.conn.commit()
        except Exception as e:
            # في حالة خطأ، عمل rollback وإعادة المحاولة
            try:
                self.conn.rollback()
                print(f"   [WARNING] خطأ في تحديث رقم الهاتف، تم التراجع: {e}")
            except:
                pass
    
    def process_employee(self, data: Dict, row_number: int):
        """معالجة بيانات موظف واحد"""
        # استخراج البيانات
        code = self.clean_text(data.get('Code'))
        emp_name = self.clean_text(data.get('Emp. Name'))
        position = self.clean_text(data.get('Position'))
        department = self.clean_text(data.get('Department'))
        ext_no = self.clean_text(data.get('Ext.NO'))
        email_string = data.get('Email')
        mobile = self.clean_text(data.get('Mobile'))
        
        # استخراج الإيميل
        email = self.extract_first_email(email_string)
        
        # البحث عن المستخدم
        user = self.find_user(email, code, emp_name)
        
        if not user:
            print(f"[SKIP] الصف {row_number}: المستخدم غير موجود - تم التخطي (لا إنشاء مستخدمين جدد)")
            self.skipped_count += 1
            return
        
        user_id = user['id']
        print(f"[UPDATE] الصف {row_number}: تحديث المستخدم {user['name']} ({email or 'لا يوجد إيميل'})")
        
        updates = {}
        updated = False
        
        # تحديث Position (job_title)
        if position and user.get('job_title') != position:
            updates['job_title'] = position
            updated = True
            print(f"   [OK] المسمى الوظيفي: {position}")
        
        # تحديث Department
        if department:
            dept = self.find_department(department)
            if dept and user.get('department_id') != dept['id']:
                updates['department_id'] = dept['id']
                updated = True
                print(f"   [OK] القسم: {dept['name']}")
            elif not dept:
                print(f"   [WARNING] القسم '{department}' غير موجود")
        
        # تحديث Avaya Extension (Ext.NO → avaya_extension)
        if ext_no:
            ext_no = re.sub(r'[^\d]', '', str(ext_no))  # إزالة أي أحرف غير رقمية
            if ext_no and user.get('avaya_extension') != ext_no:
                updates['avaya_extension'] = ext_no
                updated = True
                print(f"   [OK] رقم Avaya Extension: {ext_no}")
        
        # تحديث Work Phone Number (Mobile → phone_work)
        if mobile:
            phone_work = self.extract_first_phone(mobile)
            if phone_work:
                self.update_work_phone(user_id, mobile)
                if user.get('phone_work') != phone_work:
                    updates['phone_work'] = phone_work
                    updated = True
        
        # تحديث Employee ID (إذا لم يكن موجوداً)
        if code:
            if not user['employee_id']:
                # التحقق من عدم وجود employee_id آخر بنفس القيمة
                self.cursor.execute("""
                    SELECT id FROM users WHERE employee_id = %s AND id != %s
                """, (code, user_id))
                
                existing = self.cursor.fetchone()
                if not existing:
                    updates['employee_id'] = code
                    updated = True
                    print(f"   [OK] رقم الموظف: {code}")
            elif user.get('employee_id') != code:
                # التحقق قبل التحديث
                self.cursor.execute("""
                    SELECT id FROM users WHERE employee_id = %s AND id != %s
                """, (code, user_id))
                
                existing = self.cursor.fetchone()
                if not existing:
                    updates['employee_id'] = code
                    updated = True
                    print(f"   [OK] رقم الموظف: تم التحديث إلى {code}")
        
        # تحديث الاسم (إذا كان مختلفاً)
        if emp_name and user.get('name') != emp_name:
            updates['name'] = emp_name
            updated = True
            print(f"   [OK] الاسم: تم التحديث إلى {emp_name}")
        
        # تحديث الإيميل إذا كان مختلفاً
        if email and user.get('email') != email:
            # التحقق من عدم استخدام الإيميل من قبل مستخدم آخر
            self.cursor.execute("""
                SELECT id FROM users WHERE email = %s AND id != %s
            """, (email, user_id))
            
            existing = self.cursor.fetchone()
            if not existing:
                updates['email'] = email
                updated = True
                print(f"   [OK] الإيميل: تم التحديث إلى {email}")
        
        # تطبيق التحديثات على جدول users
        if updates:
            try:
                set_clauses = []
                values = []
                for key, value in updates.items():
                    set_clauses.append(f"{key} = %s")
                    values.append(value)
                
                values.append(user_id)
                query = f"UPDATE users SET {', '.join(set_clauses)}, updated_at = NOW() WHERE id = %s"
                self.cursor.execute(query, values)
                self.conn.commit()
                
                self.updated_count += 1
                self.success_count += 1
                print("   [OK] تم التحديث بنجاح\n")
            except Exception as e:
                # في حالة خطأ، عمل rollback
                try:
                    self.conn.rollback()
                except:
                    pass
                self.errors.append(f"الصف {row_number}: خطأ في التحديث - {str(e)}")
                self.skipped_count += 1
                print(f"   [ERROR] خطأ: {e}\n")
        else:
            print("   [INFO] لا توجد تحديثات مطلوبة\n")
        
        # إضافة جميع الإيميلات إلى جدول employee_emails
        if email_string:
            self.add_employee_emails(user_id, email_string)
    
    def import_from_excel(self, excel_file: str):
        """استيراد البيانات من ملف Excel"""
        if not os.path.exists(excel_file):
            print(f"[ERROR] ملف Excel غير موجود: {excel_file}\n")
            sys.exit(1)
        
        print(f"[FILE] قراءة ملف Excel: {excel_file}\n")
        
        try:
            # قراءة ملف Excel
            wb = load_workbook(excel_file, data_only=True)
            sheet = wb.active
            
            # قراءة العناوين من الصف الثاني (الصف 2 في Excel = index 2 في Python)
            headers = {}
            for col_idx, cell in enumerate(sheet[2], start=1):
                header = self.clean_text(cell.value)
                if header:
                    col_letter = cell.column_letter
                    headers[col_letter] = header
            
            print("[HEADERS] العناوين الموجودة:")
            for col, header in headers.items():
                print(f"   {col}: {header}")
            print()
            
            # قراءة البيانات من الصف الثالث فما بعد
            data_rows = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=False), start=3):
                row_data = {}
                has_data = False
                
                for cell in row:
                    col_letter = cell.column_letter
                    if col_letter in headers:
                        value = cell.value
                        if value:
                            has_data = True
                        # معالجة الإيميلات المتعددة
                        if headers[col_letter] == 'Email' and value:
                            row_data[headers[col_letter]] = str(value)
                        else:
                            row_data[headers[col_letter]] = self.clean_text(value) if value else None
                
                if has_data:
                    data_rows.append((row_data, row_idx))
            
            print(f"[INFO] تم قراءة {len(data_rows)} صف من البيانات\n")
            print("[START] بدء التحديث...\n")
            
            # معالجة كل صف
            for row_data, row_number in data_rows:
                self.process_employee(row_data, row_number)
            
            # عرض النتائج النهائية
            self.display_results()
            
        except Exception as e:
            print(f"[ERROR] خطأ في قراءة ملف Excel: {e}\n")
            import traceback
            traceback.print_exc()
            sys.exit(1)
    
    def display_results(self):
        """عرض النتائج النهائية"""
        print("\n" + "=" * 60)
        print("[SUMMARY] ملخص النتائج:")
        print("=" * 60)
        print(f"[OK] تم التحديث بنجاح: {self.success_count} مستخدم")
        print(f"[UPDATE] تم التحديث: {self.updated_count} مستخدم")
        print(f"[SKIP] تم التخطي: {self.skipped_count} صف")
        print(f"[ERROR] الأخطاء: {len(self.errors)}")
        
        if self.errors:
            print("\n[ERRORS] قائمة الأخطاء:")
            for error in self.errors:
                print(f"   - {error}")
        
        print()


def load_db_config_from_env():
    """تحميل إعدادات قاعدة البيانات من ملف .env"""
    config = {}
    
    # قراءة ملف .env
    env_file = '.env'
    if os.path.exists(env_file):
        with open(env_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    key = key.strip()
                    value = value.strip().strip('"').strip("'")
                    config[key] = value
    
    # إعدادات قاعدة البيانات
    db_config = {
        'host': config.get('DB_HOST', '127.0.0.1'),
        'port': int(config.get('DB_PORT', '5432')),
        'database': config.get('DB_DATABASE', 'CRM_ALL'),
        'user': config.get('DB_USERNAME', 'postgres'),
        'password': config.get('DB_PASSWORD', '')
    }
    
    return db_config


def main():
    """الدالة الرئيسية"""
    try:
        # تحميل إعدادات قاعدة البيانات
        db_config = load_db_config_from_env()
        
        # إنشاء المحدث
        updater = UsersExcelUpdater(db_config)
        
        # الاتصال بقاعدة البيانات
        updater.connect_db()
        
        # تحميل البيانات المرجعية
        updater.load_reference_data()
        
        # استيراد البيانات من ملف Excel
        excel_file = 'Copy of Employee Contact Data Oct.2025_FIXED.xlsx'
        updater.import_from_excel(excel_file)
        
        # إغلاق الاتصال
        updater.close_db()
        
    except Exception as e:
        print(f"[ERROR] خطأ عام: {e}\n")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()

